import cv2,pickle,sys,os,json,base64,getopt,shutil
import numpy as np
from copy import *
try:
    from PIL import Image
    from PIL import ImageDraw
    from PIL import ImageFont
except ImportError:
    import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as xlImage

morphclose_iternum = 15

argv = sys.argv[1:]
in_path = ''; imshow = False
try:
  opts, args = getopt.getopt(argv, '-i:-o:', ['input=', 'output=', 'imshow'])
except getopt.GetoptError:
  sys.exit(2)

for opt, arg in opts:
  if opt in ('-i', '--input'):
    in_path = arg
  elif opt in ('-o', '--output'):
    out_path = arg
  elif opt == '--imshow':
    imshow = True
if not in_path:
  print('缺少输入路径'); sys.exit(2)
if not out_path:
  print('缺少输出路径'); sys.exit(2)

x_axis = 0
y_axis = 1
def sediment(arr, axis=0):  #“沉淀”，保留备用
  sediment = arr.sum(axis=axis)
  sediment = np.where((sediment > 0), 1, 0)
  sediment = sediment.tolist()
  state = sediment[0]
  bucket = [state, 0]
  result = []
  for i in sediment:
    if i == state:
      bucket[1] += 1
    else:
      result.append(copy(bucket))
      state = i
      bucket = [state, 1]
  result.append(copy(bucket))
  return result

def acompare(a, b, error=(0.9,1.1)):  #带误差地比较a和b
  return a > b*error[0] and a < b*error[1]

def bucket_sort(bucket):  #这不是经典的“桶排序”，这只是“排序一个桶内的元素”
  assert bucket
  f = lambda zone: zone.axis
  return sorted(bucket, key=f, reverse=True)  #反排，因为竖排文字从右到左

def bucket_check(zone, bucket):
  bucket = bucket_sort(bucket)
  assert bucket
  l_width = [z.width for z in bucket]
  aw = sum(l_width)/len(l_width)
  if not acompare(zone.width, aw):
    return False
  l_top = [z.top for z in bucket]
  at = sum(l_top)/len(l_top)
  if not acompare(zone.top, at):
    return False
  first_zone = bucket[0]
  last_zone = bucket[-1]
  xdis = min(abs(first_zone.axis-zone.axis), abs(last_zone.axis-zone.axis))
  if len(bucket) == 1:
    dis = bucket[0].width*1.5
  else:
    l_dis = []
    for i in range(1,len(bucket)):
      l_dis.append(bucket[i-1].axis-bucket[i].axis)
    dis = sum(l_dis)/len(l_dis)
  if xdis >= dis*2:
    return False
  return True

def bucket2info(bucket):
  bucket = bucket_sort(bucket)
  assert bucket
  if len(bucket) == 1:
    lineinfo = [(bucket[0].width, bucket[0].width*1.5)]
    pos = (bucket[0].axis, bucket[0].top)
  else:
    lineinfo = []
    l_width = [z.width for z in bucket]
    l_dis = []
    for i in range(1,len(bucket)):
      l_dis.append(bucket[i-1].axis-bucket[i].axis)
    for i in range(len(bucket)-1):
      lineinfo.append((l_width[i], l_dis[i]))
    lineinfo.append((l_width[-1], l_width[-1]*1.5))
    l_top = [z.top for z in bucket]
    at = sum(l_top)/len(l_top)
    pos = (bucket[0].axis, at)
  return {'pos':pos, 'lines':lineinfo}

def bucket_rect(bucket):
  bucket = bucket_sort(bucket)
  assert bucket
  l_top = [z.top for z in bucket]
  t = round(sum(l_top)/len(l_top))
  l_bottom = [z.bottom for z in bucket]
  b = max(l_bottom)
  l = bucket[-1].left
  r = bucket[0].right
  return (l, t, r-l, b-t)

class zone:
    def __init__(self, rect=None, contour=None):
        if rect:
            self.rect = rect
        if type(contour) != type(None):
            self.rect = cv2.boundingRect(contour)  #寻找边界矩形
        l, t, w, h = self.rect
        self.left = l; self.right = l+w; self.top = t; self.bottom = t+h
        self.width = w; self.height = h
        self.axis = (self.left+self.right)/2
    def imgslice(self, img):
        return img[self.top:self.top+h, self.left:self.left+w]
    def rel2abs(self, point):
        return (point[0]+self.top, point[1]+self.left)
    def abs2rel(self, point):
        return (point[0]-self.top, point[1]-self.left)
    def draw(self, img, color=(255,0,0)):
        l, t, w, h = self.rect
        cv2.rectangle(img, (l,t), (l+w,t+h), color, 1)
    def __repr__(self):
        return 'zone(rect=%s)' % repr(self.rect)

if in_path == 'PIPE':
  img = pickle.loads(base64.b64decode(sys.stdin.read().encode()))
else:
  img_pil = Image.open(in_path)
  img = np.asarray(img_pil)
cv2.bitwise_not(img,img)  #反相
imgc = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 3))  #只在竖直方向闭运算
imgmor = cv2.morphologyEx(img, cv2.MORPH_CLOSE, kernel, iterations=morphclose_iternum)
contours = cv2.findContours(imgmor,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)[0]

bucket_shelf = []
for c in contours:
  z = zone(contour=c)
  z.draw(imgc)
  #分类
  newbucket = True
  for bucket in bucket_shelf:
    if bucket_check(z, bucket):
      bucket.append(z)
      newbucket = False
      break
  if newbucket:
    abucket = [z]
    bucket_shelf.append(abucket)

output = []
cv2.bitwise_not(img,img)  #反相
odir = os.path.dirname(out_path)
oname = os.path.basename(out_path).split('.')[0]
pdir = odir+'\\%s_pyxl_img'%oname
os.makedirs(pdir)
#pyxl
wb = Workbook()
ws = wb.active
ws['A1'] = '图片'
ws['B1'] = '翻译填这里（行数不得大于原文，无意义就留空）'
ws.column_dimensions['B'].width = 200
for i in range(len(bucket_shelf)):
  bucket = bucket_shelf[i]
  output.append(bucket2info(bucket))
  l, t, w, h = bucket_rect(bucket)
  cv2.rectangle(imgc, (l,t), (l+w,t+h), (0,255,0), 1)
  #保存并用pyxl读取
  imgslice = img[t:t+h, l:l+w]
  imgslice_pil = Image.fromarray(imgslice)
  pname = 'img_%d.png'%i
  ppath = odir+'\\%s_pyxl_img\\'%oname+pname
  imgslice_pil.save(ppath)
  xlimg = xlImage(ppath)
  xlimg.width, xlimg.height = imgslice_pil.size
  ws['A%d'%(i+2)] = ''
  ws.add_image(xlimg, 'A%d'%(i+2))
  if w*0.8 > ws.column_dimensions['A'].width:
    ws.column_dimensions['A'].width = w*0.3
  ws.row_dimensions[i+2].height = h*0.75+20
ws_json = wb.create_sheet('json', 1)
ws_json['A1'] = json.dumps(output)
ws_json['A2'] = '这里存放着文字位置的json数据。不要动，除非你知道你在干什么。'
wb.save(out_path)
shutil.rmtree(pdir)

if imshow:
  imgc_pil = Image.fromarray(imgc)
  imgc_pil.show()
