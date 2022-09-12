import numpy as np
import os, shutil
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as xlImage
from PIL import Image, ImageDraw, ImageFont
from lib_scripts.myio import str2obj, obj2str

alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def xlindex(x, y):
    return "%s%d" % (alphabet[x], y+1)

def getcolumnw(ws, column):
    return ws.column_dimensions[column].width / 0.3
def setcolumnw(ws, column, w):
    ws.column_dimensions[column].width = w * 0.3

def getrowh(ws, row):
    return (ws.row_dimensions[row].height - 20) / 0.75
def setrowh(ws, row, h):
    ws.row_dimensions[row].height = h * 0.75 + 20


class PyxlCtx:
    def __init__(self, wb, out_path):
        self.out_path = out_path
        self.wb = wb
        self.imgcount = 0
        outdir = os.path.dirname(out_path)
        outfilename = os.path.basename(out_path).split(".")[0]
        self.picdir = outdir + "\\%s_pyxl_img\\" % outfilename
    def __enter__(self):
        try:
            os.makedirs(self.picdir)
        except:
            print('图片缓存文件夹已存在，将直接使用此文件夹。')
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.out_path)
        shutil.rmtree(self.picdir)
    def insertcvimg(self, ws, pos, imgarr):
        self.imgcount += 1
        picname = "img_%d.png" % self.imgcount
        picpath = self.picdir + picname
        img_pil = Image.fromarray(imgarr)
        img_pil.save(picpath)
        xlimg = xlImage(picpath)
        xlimg.width, xlimg.height = img_pil.size
        ws[pos] = ""; ws.add_image(xlimg, pos)
        # 调节格大小
        w = imgarr.shape[1]
        h = imgarr.shape[0]
        if w*0.8 > getcolumnw(ws, pos[0]):
            setcolumnw(ws, pos[0], w)
        setrowh(ws, int(pos[1:]), h)


def py2xl(obj, xlpath, default_width=(), extra=()):
    # pyxl
    with PyxlCtx(Workbook(), xlpath) as ctxman:
        wb = ctxman.wb; ws = wb.active
        for i in range(len(default_width)):
            ws.column_dimensions[alphabet[i]].width = default_width[i]
        rown = -1
        for row in obj:
            rown += 1
            cvimgidx = [i for i in range(len(row)) if type(row[i])==np.ndarray]
            arow = ['' if i in cvimgidx else row[i] for i in range(len(row))]
            ws.append(arow)
            for idx in cvimgidx:
                ctxman.insertcvimg(ws, xlindex(idx, rown), row[idx])
        for ex in extra:
            name, data = ex
            data = obj2str(data)
            nws = wb.create_sheet(name)
            stridxs = (*range(0, len(data), 100), len(data))
            for n in range(len(stridxs)-1):
                nws['A%d'%(n+1)] = data[stridxs[n]:stridxs[n+1]]

def xl2py(xlpath, itername='rows', extra=()):
    wb = load_workbook(xlpath); ws = wb.active
    extradat = (str2obj(''.join(c.value for c in list(wb[name].columns)[0])) for name in extra)
    return [[y.value for y in x] for x in getattr(ws, itername)], extradat